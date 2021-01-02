#!/Users/vinodverghese/anaconda3/bin/ python3

# import time
import logging
import openpyxl as pyxl
import pandas as pd

import datetime

###############################################################################
# Excel sheet names
###############################################################################
RATESHEET = 'FX Rates'
INSHEET = 'IN Stocks'
INUTSHEET = 'IN UT'
SGSHEET = 'SG Stocks'
SGUTSHEET = 'SG UT'
USSHEET = 'US Stocks'
CRYPTOSHEET = 'Crypto'
GOLDSHEET = 'Gold'
FSSHEET = 'FS'
EODSHEET = 'EOD'

###############################################################################
# Paths
###############################################################################
SOURCE = '/Volumes/Secomba/vinodverghese/Boxcryptor/Dropbox/Personal/Finance/Financial statement Master.xlsx'


def setup_logger():
    """Logging setup (hierarchy: DIWEC)"""
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)

    formatter = logging.Formatter('%(levelname)s:%(name)s:%(asctime)s:%(message)s')

    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.INFO)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    file_handler = logging.FileHandler('stocksdebug.log')
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.DEBUG)
    logger.addHandler(file_handler)

    file_handler = logging.FileHandler('stockserr.log')
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.ERROR)
    logger.addHandler(file_handler)

    return logger


def get_date(format):
    # Get todays date
    now = datetime.datetime.now()

    # format date as dd-mm-yyyy
    if format == 1:
        dte = str(now.day) + '-' + str(now.month) + '-' + str(now.year)

    # format date as dd-mm-yyyy hh:mm:ssss
    if format == 2:
        dte = str(now.day) + '-' + str(now.month) + '-' + str(now.year) + ' ' + \
              str(now.hour) + ':' + str(now.minute) + ':' + str(now.second)

    logger.info('Todays date: %s' % dte)

    return dte


def load_fs_workbook():
    """ Load workbook """
    logger.info('Loading fs workbook')

    try:
        wb = pyxl.load_workbook(SOURCE)

    except Exception as e:
        logger.exception('Workbook load error : %s' % e)
        quit()

    logger.info('Loading of fs complete')

    return wb


def update_eod():
    """ Update EOD sheet """

    logger.info('\nUpdating EOD total..')

    # Get todays date
    dte = get_date(1)

    # Read forex rates
    rtdf = pd.read_excel(SOURCE, sheet_name='FX Rates', usecols=['Currency', 'Rate'], nrows=5)
    ussgrt = rtdf.loc[rtdf['Currency'] == 'USD to SGD', 'Rate']
    sginrt = rtdf.loc[rtdf['Currency'] == 'SGD to INR', 'Rate']

    # Read US Stock Value
    usdf = pd.read_excel(SOURCE, sheet_name='US Stocks', usecols=['Ticker', 'Current Value'])
    usdf = usdf[-usdf['Ticker'].isnull()]
    ustotal = usdf['Current Value'].sum() * ussgrt

    # Read IN Stock Value
    indf = pd.read_excel(SOURCE, sheet_name='IN Stocks', usecols=['Ticker', 'Current Value'])
    indf = indf[-indf['Ticker'].isnull()]
    intotal = indf['Current Value'].sum() / sginrt

    # Read SG Stock Value
    sgdf = pd.read_excel(SOURCE, sheet_name='SG Stocks', usecols=['Ticker', 'Current Value', 'Held'])
    sgdf = sgdf[-sgdf['Ticker'].isnull()]
    sgdf = sgdf[sgdf['Held'] == 'POEMS']
    sgtotal = sgdf['Current Value'].sum()

    # Read Crypto Value
    crydf = pd.read_excel(SOURCE, sheet_name='Crypto', usecols=['Ticker', 'Current Value'])
    crydf = crydf[-crydf['Ticker'].isnull()]
    crytotal = crydf['Current Value'].sum() * ussgrt

    # Read Gold Value
    glddf = pd.read_excel(SOURCE, sheet_name='Gold', usecols=['Ticker', 'Total'])
    glddf = glddf[-glddf['Ticker'].isnull()]
    gldtotal = glddf['Total'].sum()

    # Read IN UT Value
    inutdf = pd.read_excel(SOURCE, sheet_name='IN UT', usecols=['Fund Name', 'Value'])
    inutdf = inutdf[-inutdf['Fund Name'].isnull()]
    inuttotal = inutdf['Value'].sum() / sginrt

    # Read SG UT Value
    sgutdf = pd.read_excel(SOURCE, sheet_name='SG UT', usecols=['Fund Name', 'Value'])
    sgutdf = sgutdf[-sgutdf['Fund Name'].isnull()]
    sguttotal = sgutdf['Value'].sum()

    destws = wb[EODSHEET]
    row = 2

    while (destws.cell(row, 1).value):
        row += 1

    destws.cell(row, 1).value = dte
    destws.cell(row, 2).value = int(ustotal)
    destws.cell(row, 3).value = int(intotal)
    destws.cell(row, 4).value = int(sgtotal)
    destws.cell(row, 5).value = int(crytotal)
    destws.cell(row, 6).value = int(gldtotal)
    destws.cell(row, 7).value = int(inuttotal)
    destws.cell(row, 8).value = int(sguttotal)

    logger.info('EOD Sheet Updated')


def save_fs_workbook():
    """ save workbook """

    ws = wb[FSSHEET]

    # Get todays date
    dte = get_date(2)

    ws.cell(30, 3).value = dte

    logger.info('Saving FS workbook..')

    try:
        wb.save(SOURCE)

    except Exception as e:
        logger.exception('Workbook save error : %s' % e)
        quit()

    logger.info('Saved')


if __name__ == '__main__':
    logger = setup_logger()

    wb = load_fs_workbook()

    update_eod()

    save_fs_workbook()
