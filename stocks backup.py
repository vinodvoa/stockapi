#!/usr/bin/env python3

"""
This program gets prices for India stock, India mutual fund, US stock & Crypto
prices via free APIs / web scraping and updates a spreadsheet
Add Holiday handling - http://www.rightline.net/calendar/market-holidays.html
"""
import os
import os.path
import requests
import shutil
import time
import json
import logging
import openpyxl as pyxl

from bs4 import BeautifulSoup
from datetime import datetime, timedelta


###############################################################################
# Excel sheet name
###############################################################################
RATESHEET = 'FX Rates'
INSHEET = 'IN Stocks'
INUTSHEET = 'IN UT'
SGSHEET = 'SG Stocks'
SGUTSHEET = 'SG UT'
USSHEET = 'US Stocks'
CRYPTOSHEET = 'Crypto'
GOLDSHEET = 'Gold'
FSSHEET = 'FS'

PORTSHEET = 'Portfolio'

###############################################################################
# Paths
###############################################################################
BKUPDIR = '/Volumes/Secomba/vinodverghese/Boxcryptor/Dropbox/Personal/Finance/Backup'
SOURCE = '/Volumes/Secomba/vinodverghese/Boxcryptor/Dropbox/Personal/Finance/Financial statement Master.xlsx'
SOURCE1 = '/Volumes/Secomba/vinodverghese/Boxcryptor/Dropbox/Personal/Finance/Portfolio Balance.xlsx'

# TARGET = '/Users/vinodverghese/Dropbox/Python/Learning/Completed/Stockapi/stockapi/Financial statement Master.xlsx'

###############################################################################
# URLs
###############################################################################
#GOLDURL = 'https://www.moneymetals.com/precious-metals-charts/gold-price'
GOLDURL  = 'https://finance.yahoo.com/quote/XAUUSD=X'
SGUTURL  = 'https://www.ocbc.com/rates/daily_price_unit_trust.html'
USDFXURL = 'https://www.exchange-rates.org/currentRates/P/USD'
CADFXURL = 'https://www.exchange-rates.org/currentRates/P/CAD'
GBPFXURL = 'https://www.exchange-rates.org/currentRates/P/GBP'
SGDFXURL = 'https://www.exchange-rates.org/currentRates/P/SGD'
SGYURL   = 'https://finance.yahoo.com/quote/'

quandlkey = os.environ.get('QUANDLKEY')

DELAY = 3

# global vars
logger = None
wb = None
wb1 = None
querydte = None


def setup_logger():
    """Logging setup (hierarchy: DIWEC)"""
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)

    formatter = logging.Formatter('%(levelname)s:%(name)s:%(asctime)s:%(message)s')

    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.INFO)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    file_handler = logging.FileHandler('stocksdebug.log')
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.DEBUG)
    logger.addHandler(file_handler)

    file_handler = logging.FileHandler('stockserr.log')
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.ERROR)
    logger.addHandler(file_handler)

    return logger


def check_urls(url):
    page = requests.get(url)
    return page.status_code


def check_file_exists(filename):
    """ check if source file exists """
    if not (os.path.exists(filename)):
        logger.error('%s does not exist' % filename)
        raise SystemExit(99)
    else:
        logger.info('%s exist' % filename)
        return True

# def check_output_file_exists():
    # """ check if target file exists """
    # if not (os.path.exists(TARGET)):
    #     logger.error('%s does not exist' % TARGET)
    #     sys.exit(99)
    # else:
    #     logger.info('%s exist' % TARGET)


def backup_input_file():
    """ backup input file """
    # split source path
    path, filename = os.path.split(SOURCE)
    logger.debug('Path : %s / Filename : %s' % (path, filename))

    # split filename
    filewithoutext = os.path.splitext(filename)[0]
    logger.debug('Filename w/o ext : %s' % filewithoutext)

    # split file extension
    ext = os.path.splitext(filename)[1]
    logger.debug('File ext : %s' % ext)

    # Get todays date and format
    now = datetime.now()

    # format date as dd-mm-yyyy
    dte = str(now.day) + '-' + str(now.month) + '-' + str(now.year)
    logger.info('Todays date: %s' % dte)

    # Build up backup path
    bkupfilename = filewithoutext + ' ' + dte + ext
    logger.debug('Backup filename : %s' % bkupfilename)

    bkuppath = path + '/' + bkupfilename
    logger.debug('Backup path : %s' % bkuppath)

    # Backup source file
    logger.info('Backing up..')

    try:
        shutil.copy(SOURCE, BKUPDIR)

    except Exception as e:
        logger.exception('File copy error : %s' % e)
        quit()

    logger.info('Backup complete')

    # Rename file
    oldname = os.path.join(BKUPDIR, filename)
    logger.info('Rename from : %s' % oldname)

    newname = os.path.join(BKUPDIR, bkupfilename)
    logger.info('Rename to : %s' % newname)

    try:
        os.rename(oldname, newname)

    except Exception as e:
        logger.exception('Rename error : %s' % e)
        quit()

    logger.info('Backup renamed')


# def get_query_date():
#     """ get query date to use """

#     # Get todays date and format
#     now = datetime.now()

#     if now.weekday() in range(0, 5):  # Weekdays
#         if now.weekday() == 0:  # Mon
#             now = datetime.now() - timedelta(3)
#             querydte = str(now.year) + '-' + str(now.month) + '-' + str(now.day)
#         else:
#             querydte = str(now.year) + '-' + str(now.month) + '-' + str(now.day - 1)
#     else:
#         if now.weekday() == 5:  # Sat
#             now = datetime.now() - timedelta(1)
#             querydte = str(now.year) + '-' + str(now.month) + '-' + str(now.day)
#         else:
#             if now.weekday() == 6:  # Sun
#                 now = datetime.now() - timedelta(2)
#                 querydte = str(now.year) + '-' + str(now.month) + '-' + str(now.day)

#     logger.info('Weekday : %s' % now.weekday())
#     logger.info('Query date : %s' % querydte)

#     return querydte


def load_fs_workbook():
    """ Load workbook """
    logger.info('Loading fs workbook')

    try:
        wb = pyxl.load_workbook(SOURCE)

    except Exception as e:
        logger.exception('Workbook load error : %s' % e)
        quit()

    logger.info('Loading of fs complete')

    return wb

def load_pb_workbook():
    """ Load workbook """
    logger.info('Loading pb workbook')

    try:
        wb1 = pyxl.load_workbook(SOURCE1)

    except Exception as e:
        logger.exception('Workbook load error : %s' % e)
        quit()

    logger.info('Loading of pb complete')

    return wb1


def get_forex_rates():
    """ Get forex rates """
    logger.info('Getting forex rates..')

    ws = wb[RATESHEET]

    row = 1

    while (ws.cell(row, 1).value):
        currency = ws.cell(row, 1).value
        logger.info('Currency : %s' % currency)

        if currency[0: 3] == 'USD':
            logger.debug('USD URL : %s' % USDFXURL)
            page = requests.get(USDFXURL)
        elif currency[0: 3] == 'GBP':
            logger.debug('GBP URL : %s' % GBPFXURL)
            page = requests.get(GBPFXURL)
        elif currency[0: 3] == 'CAD':
            logger.debug('CAD URL : %s' % CADFXURL)
            page = requests.get(CADFXURL)
        elif currency[0: 3] == 'SGD':
            logger.debug('SGD URL : %s' % SGDFXURL)
            page = requests.get(SGDFXURL)

        logger.debug('Forex page : %s' % page)

        # create BS
        soup = BeautifulSoup(page.content, 'lxml')

        tr = soup.findAll('tr')
        logger.debug('Table rows : %s' % tr)

        for td in tr:
            tds = td.findAll('td')
            logger.debug('Table detail: %s' % tds)

            for moretds in tds:
                logger.debug('More table details: %s' % moretds)

                if currency == 'USD to INR':
                    if moretds.find('a', title='Indian Rupee'):
                        logger.debug('Next sibling: %s' % moretds.next_sibling)
                        usdinrrate = float(moretds.next_sibling.strong.text)
                        logger.info('USD to INR Rate : %s' % usdinrrate)
                        ws.cell(row, 2).value = usdinrrate
                        row += 1
                        break

                elif currency == 'USD to SGD':
                    if moretds.find('a', title='Singapore Dollar'):
                        logger.debug('Next sibling: %s' % moretds.next_sibling)
                        usdsgdrate = float(moretds.next_sibling.strong.text)
                        logger.info('USD to SGD Rate : %s' % usdsgdrate)
                        ws.cell(row, 2).value = usdsgdrate
                        row += 1
                        break

                elif currency == 'SGD to INR':
                    if moretds.find('a', title='Indian Rupee'):
                        logger.debug('Next sibling: %s' % moretds.next_sibling)
                        sgdinrrate = float(moretds.next_sibling.strong.text)
                        logger.info('SGD to INR Rate : %s' % sgdinrrate)
                        ws.cell(row, 2).value = sgdinrrate
                        row += 1
                        break

                elif currency == 'CAD to SGD':
                    if moretds.find('a', title='Singapore Dollar'):
                        logger.debug('Next sibling: %s' % moretds.next_sibling)
                        cadsgdrate = float(moretds.next_sibling.strong.text)
                        logger.info('CAD to SGD Rate : %s' % cadsgdrate)
                        ws.cell(row, 2).value = cadsgdrate
                        row += 1
                        break

                elif currency == 'GBP to SGD':
                    if moretds.find('a', title='Singapore Dollar'):
                        logger.debug('Next sibling: %s' % moretds.next_sibling)
                        gbpsgdrate = float(moretds.next_sibling.strong.text)
                        logger.info('GBP to SGD Rate : %s' % gbpsgdrate)
                        ws.cell(row, 2).value = gbpsgdrate
                        row += 1
                        break

def get_security_price_from_yahoo(type, country, sheetname, tickercol, srcurl, searchtag, searchclass, pricecol):
    """ Get security price """
    logger.info('/nGetting stock prices from ' + srcurl)

    ws = wb[sheetname]

    row = 2

    while (ws.cell(row, tickercol).value):
        ticker = ws.cell(row, tickercol).value

        if type == 'UT':
            if country == 'IN':
                yahoourl = srcurl + str(ticker) + '.BO?p=' + str(ticker) + '.BO&.tsrc=fin-srch-v1'
        else:    
            if country == 'IN':
                if ticker == 'SRGHFL':
                    yahoourl = srcurl + ticker + '.BO?p=' + ticker + '.BO&.tsrc=fin-srch'
                else:
                    yahoourl = srcurl + ticker + '.NS?p=' + ticker + '.NS&.tsrc=fin-srch'

            else:
                if country == 'SG' or country == 'US':
                    yahoourl = srcurl + ticker + '?p=' + ticker + '&.tsrc=fin-srch'

        logger.debug('URL : %s' % yahoourl)

        page = requests.get(yahoourl)

        soup = BeautifulSoup(page.content, 'lxml')

        try:
            divtag = soup.find(searchtag, class_=searchclass)

        except Exception as e:
            logger.exception('Find error : %s' % e)
            row += 1
            continue

        try:
            price = divtag.div.span.text

        except Exception as e:
            logger.exception('Tag not found : %s' % e)
            row += 1
            continue

        ws.cell(row, pricecol).value = float(price.replace(',', ''))
        logger.info('Ticker / Price : %s / %s' % (ticker, ws.cell(row, pricecol).value))

        row += 1
        time.sleep(DELAY)


# def get_india_stock_prices():
#     """ Get India stock prices """
#     logger.info('Getting Indian stock prices..')

#     ws = wb[INSHEET]

#     row = 2

#     while (ws.cell(row, 2).value):
#         inticker = ws.cell(row, 2).value
#         logger.info('IN Ticker : %s' % inticker)

#         if inticker == 'SRGHFL':
#             yahoourl = 'https://in.finance.yahoo.com/quote/' + inticker + \
#                 '.BO?p=' + inticker + '.BO&.tsrc=fin-srch-v1'
#         else:
#             yahoourl = 'https://in.finance.yahoo.com/quote/' + inticker + \
#                 '.NS?p=' + inticker + '.NS&.tsrc=fin-srch'

#         logger.debug('SG Stock URL : %s' % yahoourl)

#         page = requests.get(yahoourl)

#         soup = BeautifulSoup(page.content, 'lxml')

#         try:
#             divtag = soup.find('div', class_='My(6px) Pos(r) smartphone_Mt(6px)')

#         except Exception as e:
#             logger.exception('Index error : %s' % e)
#             row += 1
#             continue

#         price = divtag.div.span.text
#         ws.cell(row, 6).value = float(price.replace(',', ''))
#         logger.info('Price : %s' % ws.cell(row, 6).value)

#         row += 1
#         time.sleep(DELAY)


def get_india_ut_prices():
    """ Get India UT prices """
    logger.info('Getting Indian UT prices..')

    ws = wb[INUTSHEET]

    row = 2

    while (ws.cell(row, 2).value):
        indiautticker = ws.cell(row, 2).value
        logger.info('Ticker : %s' % ws.cell(row, 1).value)

        yahoourl = 'https://in.finance.yahoo.com/quote/' + \
            str(indiautticker) + '.BO?p=' + \
            str(indiautticker) + '.BO&.tsrc=fin-srch-v1'

        logger.debug('IN UT URL : %s' % yahoourl)

        page = requests.get(yahoourl)

        soup = BeautifulSoup(page.content, 'lxml')

        try:
            divtag = soup.find('div', class_='My(6px) Pos(r) smartphone_Mt(6px)')

        except Exception as e:
            logger.exception('Index error : %s' % e)
            row += 1
            continue

        price = divtag.div.span.text
        ws.cell(row, 12).value = float(price.replace(',', ''))
        logger.info('Price : %s' % ws.cell(row, 12).value)

        row += 1
        time.sleep(DELAY)


# def get_singapore_stock_prices():
#     """ Get Singapore stock prices """
#     logger.info('Getting Singapore stock prices..')

#     ws = wb[SGSHEET]

#     row = 2

#     while (ws.cell(row, 2).value):
#         sgticker = ws.cell(row, 2).value
#         logger.info('SG Ticker : %s' % sgticker)

#         sgtickerurl = 'https://finance.yahoo.com/quote/' + sgticker + \
#             '?p=' + sgticker + '&.tsrc=fin-srch'

#         logger.debug('SG Stock URL : %s' % sgtickerurl)

#         page = requests.get(sgtickerurl)

#         soup = BeautifulSoup(page.content, 'lxml')

#         divtag = soup.find('div', class_='My(6px) Pos(r) smartphone_Mt(6px)')
#         price = divtag.div.span.text

#         ws.cell(row, 8).value = float(price)
#         logger.info('Price : %s' % ws.cell(row, 8).value)
#         row += 1


def get_singapore_ut_prices():
    """ Get Singapore UT price """
    logger.info('Getting SG UT price..')

    ws = wb[SGUTSHEET]

    try:
        page = requests.get(SGUTURL)

    except Exception as e:
        logger.exception('URL open error : %s' % e)
        quit()

    logger.debug('HTML : %s' % page)

    soup = BeautifulSoup(page.content, 'lxml')
    logger.debug('Soup : %s' % soup.body)

    panel = soup.find('a', string='Infinity US 500 Stock Index Fund SGD')
    logger.debug('Tag : %s' % panel)

    price = panel.next_element.next_element.text

    ws['L2'].value = float(price)
    logger.info('UT Price : %s' % ws['L2'].value)


# def get_us_stock_prices():
#     """ Get US stock prices """
#     logger.info('Getting US stock prices..')

#     ws = wb[USSHEET]

#     row = 2

#     while (ws.cell(row, 2).value):
#         usticker = ws.cell(row, 2).value
#         logger.info('US stock ticker : %s' % usticker)

#         ustickerurl = 'https://finance.yahoo.com/quote/' + usticker + \
#             '?p=' + usticker + '&.tsrc=fin-srch'

#         page = requests.get(ustickerurl)

#         soup = BeautifulSoup(page.content, 'lxml')

#         divtag = soup.find('div', class_='My(6px) Pos(r) smartphone_Mt(6px)')
#         price = divtag.div.span.text

#         ws.cell(row, 9).value = float(price)
#         logger.info('Price : %s' % ws.cell(row, 9).value)

#         row += 1
#         time.sleep(DELAY)


def get_crypto_prices():
    """ Get Crypto prices """
    logger.info('Getting Crypto prices..')

    ws = wb[CRYPTOSHEET]

    row = 2

    while (ws.cell(row, 2).value):
        cryptoticker = ws.cell(row, 2).value
        # logger.info('Crypto ticker : %s' % cryptoticker)

        cryptourl = 'https://min-api.cryptocompare.com/data/price?fsym=' + \
            cryptoticker + \
            '&tsyms=BTC,USD'
        logger.debug('Crypto URL : %s' % cryptourl)

        r = requests.get(cryptourl)
        parsed_json = r.json() 

        try:
            ws.cell(row, 4).value = float(parsed_json['USD'])

        except Exception as e:
            logger.exception('Index error : %s' % e)
            continue

        logger.info('Ticker / Price : %s / %s' % (cryptoticker, ws.cell(row, 4).value))

        logger.debug('Row : %s' % row)

        row += 1
        time.sleep(DELAY)


def get_gold_price():
    logger.info('Getting Gold price..')

    ws = wb[GOLDSHEET]

    page = requests.get(GOLDURL)
    logger.debug('HTML : %s' % page)

    soup = BeautifulSoup(page.content, 'lxml')
    logger.debug('Soup : %s' % soup.body)

    try:
        panel = soup.find('div', {'class': 'My(6px) Pos(r) smartphone_Mt(6px)'})
        logger.debug('Tag : %s' % panel)
        price = panel.div.span.text.replace(',', '')
        ws['D7'].value = float(price)
        logger.info('Gold price: % s' % ws['D7'].value)

    except Exception as e:
        logger.exception('Gold price retrieve error : %s' % e)


def save_fs_workbook():
    """ save workbook """
    logger.info('Saving FS workbook..')

    try:
        wb.save(SOURCE)

    except Exception as e:
        logger.exception('Workbook save error : %s' % e)
        quit()

    logger.info('Saved')

def save_pb_workbook():
    logger.info('Saving portfolio workbook..')

    ws = wb[FSSHEET]
    ws1 = wb1[PORTSHEET]

    # Get todays date and format
    now = datetime.now()

    # format date as dd-mm-yyyy
    dte = str(now.day) + '-' + str(now.month) + '-' + str(now.year)

    writeflag = True
    row = 1
    cr = 'A' + str(row)
    
    while ws1[cr].value != None:
        if ws1[cr].value == dte:
            writeflag = False
            break
        else:
            row = row + 1
            cr = 'A' + str(row)

    if writeflag:
        ws1['A' + str(row)].value = dte
        ws1['B' + str(row)].value = ws['C7'].value     
        ws1['C' + str(row)].value = ws['C8'].value       
        ws1['D' + str(row)].value = ws['C9'].value
        ws1['E' + str(row)].value = ws['C10'].value
        ws1['F' + str(row)].value = ws['C11'].value
        ws1['G' + str(row)].value = ws['C12'].value
        ws1['H' + str(row)].value = ws['C13'].value
        ws1['I' + str(row)].value = ws['C14'].value
        ws1['J' + str(row)].value = ws['C15'].value

        try:
            wb1.save(SOURCE1)

        except Exception as e:
            logger.exception('Workbook save error : %s' % e)
            quit()

        logger.info('Portfolio Bal Saved')


if __name__ == '__main__':
    logger = setup_logger()
    check_file_exists(SOURCE)
    ## check_file_exists(SOURCE1)
    backup_input_file()

    ## querydte = get_query_date()

    wb = load_fs_workbook()
    ## wb1 = load_pb_workbook()

    get_forex_rates()

    get_security_price_from_yahoo('EQ', 'IN', INSHEET, 2, 'https://in.finance.yahoo.com/quote/','div', 'My(6px) Pos(r) smartphone_Mt(6px)', 6)
    get_security_price_from_yahoo('UT', 'IN', INUTSHEET, 2, 'https://in.finance.yahoo.com/quote/','div', 'My(6px) Pos(r) smartphone_Mt(6px)', 12)
    # get_india_stock_prices()
    # get_india_ut_prices()

    get_security_price_from_yahoo('EQ', 'SG', SGSHEET, 2, 'https://sg.finance.yahoo.com/quote/','div', 'My(6px) Pos(r) smartphone_Mt(6px)', 8)
    # get_singapore_stock_prices()
    get_singapore_ut_prices()

    get_security_price_from_yahoo('EQ', 'US', USSHEET, 2, 'https://finance.yahoo.com/quote/','div', 'My(6px) Pos(r) smartphone_Mt(6px)', 9)
    # get_us_stock_prices()

    get_crypto_prices()
    get_gold_price()

    save_fs_workbook()
    ## save_pb_workbook()