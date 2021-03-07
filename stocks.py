#!/Users/vinodverghese/anaconda3/bin/ python3

"""
This program gets prices for India stocks, India mutual fund, US stocks & Crypto
prices via free APIs / web scraping and updates a spreadsheet
"""
import os
import os.path
import requests
import shutil
import time
import logging
import openpyxl as pyxl
import datetime
import sqlite3

from bs4 import BeautifulSoup

###############################################################################
# Excel sheet names
###############################################################################
RATESHEET = 'FX Rates'
INSHEET = 'IN Stocks'
INWLSHEET = 'IN WL'
INUTSHEET = 'IN UT'
SGSHEET = 'SG Stocks'
SGUTSHEET = 'SG UT'
USSHEET = 'US Stocks'
USWLSHEET = 'US WL'
CRYPTOSHEET = 'Crypto'
GOLDSHEET = 'Gold'
FSSHEET = 'FS'
EODSHEET = 'EOD'

###############################################################################
# Paths
###############################################################################
BKUPDIR = '/Volumes/Secomba/vinodverghese/Boxcryptor/Dropbox/Personal/Finance/Backup'
SOURCE = '/Volumes/Secomba/vinodverghese/Boxcryptor/Dropbox/Personal/Finance/Financial statement Master.xlsx'


###############################################################################
# URLs
###############################################################################
USDFXURL = 'https://www.exchange-rates.org/currentRates/P/USD'
CADFXURL = 'https://www.exchange-rates.org/currentRates/P/CAD'
GBPFXURL = 'https://www.exchange-rates.org/currentRates/P/GBP'
SGDFXURL = 'https://www.exchange-rates.org/currentRates/P/SGD'

# global vars
wb = None
logger = None
fair_pat_missing = None

DELAY = 1
ETF = ['IVV', 'VFH', 'MCHI', 'VHT', 'SLYG', 'XLE', 'DVY', 'VOE', 'LIT']


def setup_logger():
    """Logging setup (hierarchy: DIWEC)"""
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)

    formatter = logging.Formatter('%(levelname)s:%(name)s:%(asctime)s:%(message)s')

    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.INFO)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    file_handler = logging.FileHandler('stocksdebug.log')
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.DEBUG)
    logger.addHandler(file_handler)

    file_handler = logging.FileHandler('stockserr.log')
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.ERROR)
    logger.addHandler(file_handler)

    return logger


def setup_database():

    try:
        # conn = sqlite3.connect(':memory:')
        conn = sqlite3.connect('usstocks.db')

    except Exception as e:
        logger.info('Database connection error : %s' % e)
        quit()

    try:
        cur = conn.cursor()

    except Exception as e:
        logger.error('Cursor error : %s' % e)
        quit()

    try:
        cur.execute("""
                    CREATE TABLE usratios(
                        date TEXT NOT NULL,
                        ticker TEXT NOT NULL,
                        sector TEXT,
                        industry TEXT,
                        curr_price REAL,
                        us_mcap TEXT,
                        us_ev TEXT,
                        us_ttm_pe REAL,
                        us_fwd_pe REAL,
                        us_peg REAL,
                        us_ttm_ps REAL,
                        us_pb REAL,
                        us_ev_rev REAL,
                        us_ev_ebitda REAL,
                        us_debt_equity REAL,
                        us_fcf TEXT
                )""")
                        # PRIMARY KEY (date, ticker)


    except Exception as e:
        logger.error('Table create error : %s' % e)
        pass

    else:
        logger.info('Database created')

    try:
        conn.commit()

    except Exception as e:
        logger.error('Commit error : %s' % e)

    else:
        conn.close
        return conn, cur


def check_urls(url):
    page = requests.get(url)
    return page.status_code


def check_file_exists(filename):
    """ check if source file exists """
    if not (os.path.exists(filename)):
        logger.error('%s does not exist' % filename)
        raise SystemExit(99)
    else:
        logger.info('%s exist' % filename)
        return True


def get_date(format):
    # Get todays date
    now = datetime.datetime.now()

    # format date as dd-mm-yyyy
    if format == 1:
        dte = str(now.day) + '-' + str(now.month) + '-' + str(now.year)

    # format date as dd-mm-yyyy hh:mm:ssss
    if format == 2:
        dte = str(now.day) + '-' + str(now.month) + '-' + str(now.year) + ' ' + \
              str(now.hour) + ':' + str(now.minute) + ':' + str(now.second)

    logger.info('Todays date: %s' % dte)

    return dte


def backup_input_file():
    """ backup input file """
    # split source path
    path, filename = os.path.split(SOURCE)
    logger.debug('Path : %s / Filename : %s' % (path, filename))

    # split filename
    filewithoutext = os.path.splitext(filename)[0]
    logger.debug('Filename w/o ext : %s' % filewithoutext)

    # split file extension
    ext = os.path.splitext(filename)[1]
    logger.debug('File ext : %s' % ext)

    # Get todays date
    dte = get_date(1)

    # Build up backup path
    bkupfilename = filewithoutext + ' ' + dte + ext
    logger.debug('Backup filename : %s' % bkupfilename)

    bkuppath = path + '/' + bkupfilename
    logger.debug('Backup path : %s' % bkuppath)

    # Backup source file
    logger.info('Backing up..')

    try:
        shutil.copy(SOURCE, BKUPDIR)

    except Exception as e:
        logger.exception('File copy error : %s' % e)
        quit()

    logger.info('Backup complete')

    # Rename file
    oldname = os.path.join(BKUPDIR, filename)
    logger.info('Rename from : %s' % oldname)

    newname = os.path.join(BKUPDIR, bkupfilename)
    logger.info('Rename to : %s' % newname)

    try:
        os.rename(oldname, newname)

    except Exception as e:
        logger.exception('Rename error : %s' % e)
        quit()

    logger.info('Backup renamed')


def load_fs_workbook():
    """ Load workbook """
    logger.info('Loading fs workbook')

    try:
        wb = pyxl.load_workbook(filename=SOURCE)

    except Exception as e:
        logger.exception('Workbook load error : %s' % e)
        quit()

    logger.info('Loading of fs complete')

    return wb


def get_forex_rates():
    """ Get forex rates """
    logger.info('\nGetting forex rates..')

    ws = wb[RATESHEET]

    row = 2

    while (ws.cell(row, 1).value):
        currency = ws.cell(row, 1).value
        # logger.info('Currency : %s' % currency)

        if currency[0: 3] == 'USD':
            logger.debug('USD URL : %s' % USDFXURL)
            page = requests.get(USDFXURL)
        elif currency[0: 3] == 'GBP':
            logger.debug('GBP URL : %s' % GBPFXURL)
            page = requests.get(GBPFXURL)
        elif currency[0: 3] == 'CAD':
            logger.debug('CAD URL : %s' % CADFXURL)
            page = requests.get(CADFXURL)
        elif currency[0: 3] == 'SGD':
            logger.debug('SGD URL : %s' % SGDFXURL)
            page = requests.get(SGDFXURL)

        logger.debug('Forex page : %s' % page)

        # create BS
        soup = BeautifulSoup(page.content, 'lxml')

        tr = soup.findAll('tr')
        logger.debug('Table rows : %s' % tr)

        for td in tr:
            tds = td.findAll('td')
            logger.debug('Table detail: %s' % tds)

            for moretds in tds:
                logger.debug('More table details: %s' % moretds)

                if currency == 'USD to INR':
                    if moretds.find('a', title='Indian Rupee'):
                        logger.debug('Next sibling: %s' % moretds.next_sibling)
                        usdinrrate = float(moretds.next_sibling.strong.text)
                        logger.info('USD to INR Rate : %s' % usdinrrate)
                        ws.cell(row, 2).value = usdinrrate
                        row += 1
                        break

                elif currency == 'USD to SGD':
                    if moretds.find('a', title='Singapore Dollar'):
                        logger.debug('Next sibling: %s' % moretds.next_sibling)
                        usdsgdrate = float(moretds.next_sibling.strong.text)
                        logger.info('USD to SGD Rate : %s' % usdsgdrate)
                        ws.cell(row, 2).value = usdsgdrate
                        row += 1
                        break

                elif currency == 'SGD to INR':
                    if moretds.find('a', title='Indian Rupee'):
                        logger.debug('Next sibling: %s' % moretds.next_sibling)
                        sgdinrrate = float(moretds.next_sibling.strong.text)
                        logger.info('SGD to INR Rate : %s' % sgdinrrate)
                        ws.cell(row, 2).value = sgdinrrate
                        row += 1
                        break

                elif currency == 'CAD to SGD':
                    if moretds.find('a', title='Singapore Dollar'):
                        logger.debug('Next sibling: %s' % moretds.next_sibling)
                        cadsgdrate = float(moretds.next_sibling.strong.text)
                        logger.info('CAD to SGD Rate : %s' % cadsgdrate)
                        ws.cell(row, 2).value = cadsgdrate
                        row += 1
                        break

                elif currency == 'GBP to SGD':
                    if moretds.find('a', title='Singapore Dollar'):
                        logger.debug('Next sibling: %s' % moretds.next_sibling)
                        gbpsgdrate = float(moretds.next_sibling.strong.text)
                        logger.info('GBP to SGD Rate : %s' % gbpsgdrate)
                        ws.cell(row, 2).value = gbpsgdrate
                        row += 1
                        break


def get_security_price_from_yahoo(type, country, sheetname, tickercol, pricecol):
    """
    Get security price from Yahoo Finance
    Parameters:
        type - string; security type (EQ - Equity, UT - Unit Trust, CO - Commodities)
        country - string; listed security country (US - USA, SG - Singapore, IN - India)
        sheetname - string; input Excel sheet name
        tickercol - numeric; column in Excel sheet containing ticker
        pricecol - numeric; Excel column where ticker price is to be output
    """
    ws = wb[sheetname]

    row = 2

    # Assign url based on country
    if country == 'IN':
        srcurl = 'https://in.finance.yahoo.com/quote/'
    elif country == 'SG':
        srcurl = 'https://sg.finance.yahoo.com/quote/'
    else:
        srcurl = 'https://finance.yahoo.com/quote/'

    logger.info('\nGetting stock prices from ' + srcurl)

    fair_pat_missing = False
    fair_pat_count = 0

    # Loop through the ticker sheet column, extract the price and assign to price column
    while (ws.cell(row, tickercol).value):
        ticker = ws.cell(row, tickercol).value

        if type == 'UT':
            if country == 'IN':
                yahoourl = srcurl + str(ticker) + '.BO?p=' + str(ticker) + '.BO&.tsrc=fin-srch-v1'
            elif country == 'SG':
                yahoourl = srcurl + str(ticker) + '.SI'
        elif type == 'CO':
            yahoourl = srcurl + 'GC=F'
        else:
            if country == 'IN':
                if ticker == 'SRGHFL' or ticker == ' 500193':
                    yahoourl = srcurl + ticker + '.BO?p=' + ticker + '.BO&.tsrc=fin-srch'
                else:
                    yahoourl = srcurl + ticker + '.NS?p=' + ticker + '.NS&.tsrc=fin-srch'

            else:
                if country == 'SG' or country == 'US':
                    yahoourl = srcurl + ticker + '?p=' + ticker

        logger.debug('URL : %s' % yahoourl)

        page = requests.get(yahoourl)

        soup = BeautifulSoup(page.content, 'lxml')

        # Price
        logger.debug('Price Section')

        try:
            divtag = soup.find('div', class_='My(6px) Pos(r) smartphone_Mt(6px)')

        except Exception as e:
            logger.exception('\nFind error : %s' % e)
            pass

        try:
            price = divtag.div.span.text

        except Exception as e:
            logger.exception('\nTag not found : %s' % e)
            curr_price = 0
            pass

        try:
            curr_price = float(price.replace(',', ''))
            ws.cell(row, pricecol).value = float(price.replace(',', ''))
            logger.info('%s / %s' % (ticker, ws.cell(row, pricecol).value))

        except Exception as e:
            logger.exception('\nPrice format error : %s' % e)
            curr_price = 0
            pass

        if (country == 'US') and (ticker not in ETF) and (fair_pat_missing is False):
            # Fair Value
            logger.debug('Fair Value Section')

            try:
                valuediv = soup.findAll('div', class_='Fw(b) Fl(end)--m Fz(s) C($primaryColor')
                fair_value = valuediv[0].text
                ws.cell(row, 26).value = valuediv[0].text

            except Exception as e:
                logger.exception('\nFind error : %s' % e)
                fair_pat_count += 1

                if fair_pat_count >= 3:
                    fair_pat_missing = True
                    fair_value = ''
                pass

            # Pattern
            logger.debug('Pattern Section')

            try:
                patdiv = soup.find('div', class_='W(1/4)--mobp W(1/2) IbBox')

            except Exception as e:
                logger.exception('\nFind error : %s' % e)
                fair_pat_count += 1

                if fair_pat_count >= 3:
                    fair_pat_missing = True

                fair_pattern = ''
                pass

            try:
                fair_pattern = patdiv.div.span.text
                ws.cell(row, 27).value = patdiv.div.span.text

            except Exception as e:
                logger.exception('\nIndex error : %s' % e)
                fair_pattern = ''
                pass

        time.sleep(DELAY)

        if country == 'US' and (ticker not in ETF):
            # Stats
            logger.debug('Stats Section')

            if ticker not in ETF:
                yahoourl = srcurl + ticker + '/key-statistics?p=' + ticker

                logger.debug('URL : %s' % yahoourl)

                page = requests.get(yahoourl)

                soup = BeautifulSoup(page.content, 'lxml')

                try:
                    tds = soup.findAll('td', class_='Fw(500) Ta(end) Pstart(10px) Miw(60px)')

                except Exception as e:
                    logger.exception('\nFind error : %s' % e)
                    us_mcap = '0'
                    us_ev = 0
                    us_ttm_pe = 0
                    us_fwd_pe = 0
                    us_peg = 0
                    us_ttm_ps = 0
                    us_pb = 0
                    us_ev_rev = 0
                    us_ev_ebitda = 0
                    us_debt_equity = 0
                    us_fcf = 0
                    pass

                try:
                    us_mcap = tds[0].text
                    ws.cell(row, 15).value = tds[0].text

                    us_ev = tds[1].text
                    ws.cell(row, 16).value = tds[1].text

                    us_ttm_pe = float(tds[2].text.replace(',', '').replace('N/A', '0'))
                    ws.cell(row, 17).value = float(tds[2].text.replace(',', '').replace('N/A', '0'))

                    us_fwd_pe = float(tds[3].text.replace(',', '').replace('N/A', '0'))
                    ws.cell(row, 18).value = float(tds[3].text.replace(',', '').replace('N/A', '0'))

                    us_peg = float(tds[4].text.replace(',', '').replace('N/A', '0'))
                    ws.cell(row, 19).value = float(tds[4].text.replace(',', '').replace('N/A', '0'))

                    us_ttm_ps = float(tds[5].text.replace(',', '').replace('N/A', '0'))
                    ws.cell(row, 20).value = float(tds[5].text.replace(',', '').replace('N/A', '0'))

                    us_pb = float(tds[6].text.replace(',', '').replace('N/A', '0'))
                    ws.cell(row, 21).value = float(tds[6].text.replace(',', '').replace('N/A', '0'))

                    us_ev_rev = float(tds[7].text.replace(',', '').replace('N/A', '0'))
                    ws.cell(row, 22).value = float(tds[7].text.replace(',', '').replace('N/A', '0'))

                    us_ev_ebitda = float(tds[8].text.replace(',', '').replace('N/A', '0'))
                    ws.cell(row, 23).value = float(tds[8].text.replace(',', '').replace('N/A', '0'))

                    us_debt_equity = float(tds[54].text.replace(',', '').replace('N/A', '0'))
                    ws.cell(row, 24).value = float(tds[54].text.replace(',', '').replace('N/A', '0'))  # total debt/equity (mrq)

                    us_fcf = tds[57].text
                    ws.cell(row, 25).value = tds[57].text  # operating cashflow (ttm)

                except Exception as e:
                    logger.exception('\nStats index error : %s' % e)
                    us_mcap = '0'
                    us_ev = 0
                    us_ttm_pe = 0
                    us_fwd_pe = 0
                    us_peg = 0
                    us_ttm_ps = 0
                    us_pb = 0
                    us_ev_rev = 0
                    us_ev_ebitda = 0
                    us_debt_equity = 0
                    us_fcf = 0
                    pass

            time.sleep(DELAY)

            us_sector = ''
            us_industry = ''

            # Sector/Industry
            if (ws.cell(row, 5).value is None) or (ws.cell(row, 6).value is None):
                logger.debug('Sector/Industry Section')

                yahoourl = srcurl + ticker + '/profile?p=' + ticker

                logger.debug('URL : %s' % yahoourl)

                page = requests.get(yahoourl)

                soup = BeautifulSoup(page.content, 'lxml')

                try:
                    tds = soup.findAll('span', class_='Fw(600)')

                except Exception as e:
                    logger.exception('\nFind error : %s' % e)
                    us_sector = ''
                    us_industry = ''
                    pass

                try:
                    us_sector = tds[0].text
                    ws.cell(row, 5).value = tds[0].text

                    us_industry = tds[1].text
                    ws.cell(row, 6).value = tds[1].text

                except Exception as e:
                    logger.exception('\nFind error : %s' % e)
                    us_sector = ''
                    us_industry = ''
                    pass
          
            # with conn:
            #     cur.execute("""
            #                 INSERT INTO usratios VALUES
            #                     (
            #                      :date,
            #                      :ticker,
            #                      :sector,
            #                      :industry,
            #                      :curr_price,
            #                      :us_mcap,
            #                      :us_ev,
            #                      :us_ttm_pe,
            #                      :us_fwd_pe,
            #                      :us_peg,
            #                      :us_ttm_ps,
            #                      :us_pb,
            #                      :us_ev_rev,
            #                      :us_ev_ebitda,
            #                      :us_debt_equity,
            #                      :us_fcf
            #                     )
            #                 """, {'date': dte,
            #                       'ticker': ticker,
            #                       'sector': us_sector,
            #                       'industry': us_industry,
            #                       'curr_price': curr_price,
            #                       'us_mcap': us_mcap,
            #                       'us_ev': us_ev,
            #                       'us_ttm_pe': us_ttm_pe,
            #                       'us_fwd_pe': us_fwd_pe,
            #                       'us_peg': us_peg,
            #                       'us_ttm_ps': us_ttm_ps,
            #                       'us_pb': us_pb,
            #                       'us_ev_rev': us_ev_rev,
            #                       'us_ev_ebitda': us_ev_ebitda,
            #                       'us_debt_equity': us_debt_equity,
            #                       'us_fcf': us_fcf}
            #                 )

        elif country == 'IN' and type == 'EQ':
            # Stats
            logger.debug('IN Stats Section')

            yahoourl = srcurl + ticker + '.NS' + '/key-statistics?p=' + ticker + '.NS'

            logger.debug('URL : %s' % yahoourl)

            page = requests.get(yahoourl)

            soup = BeautifulSoup(page.content, 'lxml')

            try:
                tds = soup.findAll('td', class_='Fw(500) Ta(end) Pstart(10px) Miw(60px)')

            except Exception as e:
                logger.exception('\nFind error : %s' % e)
                pass

            try:
                ws.cell(row, 14).value = tds[0].text
                ws.cell(row, 15).value = tds[1].text
                ws.cell(row, 16).value = float(tds[2].text.replace(',', '').replace('N/A', '0'))
                ws.cell(row, 17).value = float(tds[3].text.replace(',', '').replace('N/A', '0'))
                ws.cell(row, 18).value = float(tds[4].text.replace(',', '').replace('N/A', '0'))
                ws.cell(row, 19).value = float(tds[5].text.replace(',', '').replace('N/A', '0'))
                ws.cell(row, 20).value = float(tds[6].text.replace(',', '').replace('N/A', '0'))
                ws.cell(row, 21).value = float(tds[7].text.replace(',', '').replace('N/A', '0'))
                ws.cell(row, 22).value = float(tds[8].text.replace(',', '').replace('N/A', '0'))
                ws.cell(row, 23).value = float(tds[54].text.replace(',', '').replace('N/A', '0'))  # total debt/equity (mrq)
                ws.cell(row, 24).value = tds[57].text  # operating cashflow (ttm)

            except Exception as e:
                logger.exception('\nStats index error : %s' % e)
                pass

            time.sleep(DELAY)

            # Sector/Industry
            if (ws.cell(row, 2).value is None) or (ws.cell(row, 3).value is None):
                logger.debug('IN Sector/Industry Section')

                yahoourl = srcurl + ticker + '.NS' + '/profile?p=' + ticker + '.NS'

                logger.debug('URL : %s' % yahoourl)

                page = requests.get(yahoourl)

                soup = BeautifulSoup(page.content, 'lxml')

                try:
                    tds = soup.findAll('span', class_='Fw(600)')

                except Exception as e:
                    logger.exception('\nFind error : %s' % e)
                    pass

                try:
                    ws.cell(row, 2).value = tds[0].text
                    ws.cell(row, 3).value = tds[1].text

                except Exception as e:
                    logger.exception('\nFind error : %s' % e)
                    pass

        row += 1
        time.sleep(DELAY)

        # if row == 25 or row == 50 or row == 75 or row == 100 or row == 125 or row == 150:
        #     save_fs_workbook()


def get_crypto_prices():
    """ Get Crypto prices """
    logger.info('\nGetting Crypto prices..')

    ws = wb[CRYPTOSHEET]

    row = 2

    while (ws.cell(row, 2).value):
        cryptoticker = ws.cell(row, 2).value

        cryptourl = 'https://min-api.cryptocompare.com/data/price?fsym=' + \
            cryptoticker + \
            '&tsyms=BTC,USD'
        logger.debug('Crypto URL : %s' % cryptourl)

        r = requests.get(cryptourl)
        parsed_json = r.json()

        try:
            ws.cell(row, 4).value = float(parsed_json['USD'])

        except Exception as e:
            logger.exception('Index error : %s' % e)
            pass

        logger.info('Ticker / Price : %s / %s' % (cryptoticker, ws.cell(row, 4).value))

        logger.debug('Row : %s' % row)

        row += 1
        time.sleep(DELAY)


def save_fs_workbook():
    """ save workbook """

    ws = wb[FSSHEET]

    # # Get todays date
    # dte = get_date(2)

    # ws.cell(30, 3).value = dte

    logger.info('Saving FS workbook..')

    try:
        wb.save(SOURCE)

    except Exception as e:
        logger.exception('Workbook save error : %s' % e)
        quit()

    logger.info('Saved')


if __name__ == '__main__':
    logger = setup_logger()

    check_file_exists(SOURCE)

    backup_input_file()
    # dte = get_date(2)

    wb = load_fs_workbook()
    # conn, cur = setup_database()

    get_forex_rates()
    get_crypto_prices()

    get_security_price_from_yahoo('EQ', 'US', USSHEET, 2, 9)
    get_security_price_from_yahoo('EQ', 'US', USWLSHEET, 2, 9)

    get_security_price_from_yahoo('EQ', 'IN', INSHEET, 2, 7)
    get_security_price_from_yahoo('EQ', 'IN', INWLSHEET, 1, 4)
    get_security_price_from_yahoo('UT', 'IN', INUTSHEET, 2, 12)

    get_security_price_from_yahoo('EQ', 'SG', SGSHEET, 2, 8)
    get_security_price_from_yahoo('UT', 'SG', SGUTSHEET, 2, 12)

    save_fs_workbook()

    # cur.execute("SELECT * FROM usratios")
    # print(cur.fetchall())

